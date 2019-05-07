#! python3
# cellInverter.py
# Author: Kene Udeh
# Source: Automate the Boring stuff with python Ch. 12 Project


import openpyxl


def invertCells(filename):
    """inverts all cells in a workbook
    Args:
        filename (str): excel file to invert cells in
    Returns:
        None
    """
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    newSheet = wb.create_sheet(index=0, title='inverted')

    for rowObj in sheet.rows:
        for cellObj in rowObj:
            colIndex = cellObj.column
            rowIndex = cellObj.row

            newSheet.cell(row=colIndex, column=rowIndex).value = cellObj.value

    wb.save('result_'+filename)


if __name__ == "__main__":
    invertCells('example.xlsx')