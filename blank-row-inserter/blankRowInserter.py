#! python3
# blankRowInserter.py
# Author: Kene Udeh
# Source: Automate the Boring stuff with python Ch. 12 Project

import sys


import openpyxl


def blankRowInserter(index, num_blanks, filename):
    """
    Args:
        index (int): row in file to start insert
        num_blanks (int): number of blank rows to insert
        filename (str): filename to insert blanks
    Returns:
        None
    """
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    rows = tuple(sheet.rows)

    
    for rowObj in rows[::-1]:
        for cellObj in rowObj:
            c = cellObj.column
            r = cellObj.row

            if r >= index and r < index+num_blanks:
                sheet.cell(row=r+num_blanks, column=c).value = cellObj.value
                sheet.cell(row=r, column=c).value = ''
            elif r >= index+num_blanks:
                sheet.cell(row=r+num_blanks, column=c).value = cellObj.value
    
    wb.save('result_'+filename)


if __name__ == "__main__":
    
    num_args = len(sys.argv)

    if num_args < 4:
        print("usage: python blankRowInserter.py 3 2 myProduce.xlsx")
    else:
        index = int(sys.argv[1])
        num_blanks = int(sys.argv[2])
        filename = sys.argv[3]

        blankRowInserter(index, num_blanks, filename)

    